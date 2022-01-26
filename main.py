from openpyxl import load_workbook
import creat_dict
import IndustryJudgment
import DefaultRange
import ReturnValue
import os


def main():
    filename = input('请输入将生成的文件名（默认为result):')
    if len(filename) == 0:
        filename = 'result'
    wb = load_workbook('template.xlsx')
    ws = wb.active
    main_dict_list = creat_dict.dict(ws.rows)
    cell_row_num = 1
    for single_dict in main_dict_list:
        cell_row_num = cell_row_num + 1
        # 获取企业所涉及行业名称
        name = IndustryJudgment.code_judgment(single_dict['industry_code'])
        # 获取行业对应的划型范围
        range_dict = DefaultRange.industry_range(name)
        # 获取规模
        group_size = ReturnValue.range_value(name, single_dict, range_dict)
        # 将规模写入单元格
        ws.cell(row=cell_row_num, column=7, value=group_size)
        # 生成结果文件
        wb.save(f"./result/{filename}.xlsx")
    print('已完成企业划型', f"结果以保存至./result/{filename}.xlsx")
    os.system('pause')


if __name__ == '__main__':
    main()
